#!/usr/bin/env python3
"""
Reads the AFrEnProPo listening-club spreadsheet.

Two jobs:

  current / check   Which artist is the club listening to right now, and do we
                    already have a discography report for them?
  stats             Recompute every club statistic from the raw ratings and
                    write data/ratings.json for stats.html to render.

The current artist is whichever row is filled bright green (#00FF00) by hand.
Google's xlsx export preserves manual cell fills, so this needs no API key and
no service account -- only a link-viewable spreadsheet.

Statistics are always recomputed from the raw per-member scores on Sheet1,
never scraped from the sheet's own derived tabs. Those tabs carry #REF! errors
and would poison the stats page silently.
"""

from __future__ import annotations

import argparse
import io
import json
import math
import os
import re
import sys
import unicodedata
import urllib.request
import warnings
from pathlib import Path

warnings.filterwarnings("ignore", module="openpyxl")
import openpyxl

SCRIPTS = Path(__file__).resolve().parent
REPO = SCRIPTS.parent

ALIAS_FILE = SCRIPTS / "artists.json"
REPORT_DIR = REPO / "html"
DATA_FILE = REPO / "data" / "ratings.json"

GREEN = "FF00FF00"
SESSION_TAB = "Sheet1"
EXPORT_URL = "https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"


class SheetError(RuntimeError):
    """The spreadsheet is not shaped the way we expect.

    `kind` lets callers tell a routine situation (no artist marked yet) from a
    genuine mistake (two artists marked at once) without matching on prose.
    """

    def __init__(self, message, kind=None, rows=None):
        super().__init__(message)
        self.kind = kind
        self.rows = rows or []


# --------------------------------------------------------------------------
# Loading
# --------------------------------------------------------------------------

def load_workbook(sheet_id: str | None = None, path: str | None = None):
    """Load the club workbook from a local file or Google's xlsx export.

    data_only=True gives cached formula values; cell fills survive it, so one
    load covers both the green-row lookup and the ratings.
    """
    if path:
        return openpyxl.load_workbook(path, data_only=True)
    if not sheet_id:
        raise SheetError(
            "No spreadsheet given. Pass --file for a local .xlsx, or set "
            "CLUB_SHEET_ID (or pass --sheet-id) for the Google export."
        )
    url = EXPORT_URL.format(sheet_id=sheet_id)
    try:
        with urllib.request.urlopen(url, timeout=60) as resp:
            blob = resp.read()
    except Exception as exc:
        raise SheetError(f"Could not download the spreadsheet: {exc}") from exc
    if not blob.startswith(b"PK"):
        raise SheetError(
            "Google did not return an .xlsx file. The sheet is probably no "
            "longer shared as 'anyone with the link can view'."
        )
    return openpyxl.load_workbook(io.BytesIO(blob), data_only=True)


def sessions_tab(wb):
    if SESSION_TAB not in wb.sheetnames:
        raise SheetError(f"No {SESSION_TAB!r} tab. Found: {wb.sheetnames}")
    return wb[SESSION_TAB]


# --------------------------------------------------------------------------
# Sheet shape
#
# Nothing here hardcodes a row or column position. The header row moves, the
# submitter blocks are non-contiguous, and rows are not in date order, so we
# locate the header by content and address every column by name.
# --------------------------------------------------------------------------

def find_header(ws) -> tuple[int, dict[str, int]]:
    """Return (header_row, {lowercased column name: column index})."""
    for row in range(1, 11):
        names = {}
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row, col).value
            if value is not None:
                names[str(value).strip().lower()] = col
        if "artist" in names and "album 1" in names:
            return row, names
    raise SheetError(
        "Could not find the header row -- no row in the first 10 has both "
        "'Artist' and 'Album 1'."
    )


def find_members(header: dict[str, int]) -> list[str]:
    """Rating columns are the ones between 'Special Notes' and 'Average'.

    Derived rather than hardcoded so a fourth member joining the club needs no
    code change.
    """
    left, right = header.get("special notes"), header.get("average")
    if left is None or right is None:
        raise SheetError("Header is missing 'Special Notes' or 'Average'.")
    members = sorted(
        ((col, name) for name, col in header.items() if left < col < right),
        key=lambda pair: pair[0],
    )
    if not members:
        raise SheetError("No rating columns found between Special Notes and Average.")
    return [name.title() for _, name in members]


def cell_is_green(cell) -> bool:
    fill = cell.fill
    if not fill or fill.patternType != "solid":
        return False
    return str(getattr(fill.start_color, "rgb", "")) == GREEN


# --------------------------------------------------------------------------
# Names
# --------------------------------------------------------------------------

def load_aliases() -> dict:
    with open(ALIAS_FILE, encoding="utf-8") as fh:
        return json.load(fh).get("aliases", {})


def slugify(name: str) -> str:
    """Artist name -> report filename stem, matching the existing html/ files.

    Drops parentheticals ('Bela Fleck (non-Flecktones)'), strips accents, and
    removes a leading 'The' -- the repo has mars_volta_discography.html, not
    the_mars_volta_discography.html.
    """
    text = unicodedata.normalize("NFKD", str(name))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"\([^)]*\)", " ", text)
    text = text.lower().replace("&", " and ")
    text = re.sub(r"[^a-z0-9]+", " ", text).strip()
    text = re.sub(r"^the\s+", "", text)
    return re.sub(r"\s+", "_", text)


def resolve(name: str, aliases: dict) -> tuple[str, str]:
    """Return (slug, display name) for an artist as typed in the sheet."""
    raw = str(name)
    entry = aliases.get(raw) or aliases.get(raw.strip())
    if entry:
        return entry["slug"], entry.get("display", raw.strip())
    return slugify(raw), re.sub(r"\s+", " ", raw).strip()


def report_path(slug: str) -> Path:
    return REPORT_DIR / f"{slug}_discography.html"


# --------------------------------------------------------------------------
# Reading sessions
# --------------------------------------------------------------------------

def _num(value):
    return float(value) if isinstance(value, (int, float)) else None


def _date(value):
    return value.strftime("%Y-%m-%d") if hasattr(value, "strftime") else None


def read_sessions(ws, header, header_row, members, aliases) -> list[dict]:
    """Every row that names an artist, rated or not."""
    col = header.__getitem__
    albums = [header[f"album {n}"] for n in range(1, 6) if f"album {n}" in header]
    out = []
    for row in range(header_row + 1, ws.max_row + 1):
        artist = ws.cell(row, col("artist")).value
        if artist is None or not str(artist).strip():
            continue
        slug, display = resolve(artist, aliases)
        scores = {m: _num(ws.cell(row, header[m.lower()]).value) for m in members}
        rated = {m: s for m, s in scores.items() if s is not None}
        out.append({
            "row": row,
            "artist_raw": str(artist),
            "artist": display,
            "slug": slug,
            "submitter": (str(ws.cell(row, col("submitter")).value).strip()
                          if ws.cell(row, col("submitter")).value else None),
            "genre": (str(ws.cell(row, col("genre")).value).strip()
                      if "genre" in header and ws.cell(row, col("genre")).value else None),
            "gender": (str(ws.cell(row, col("gender")).value).strip()
                       if "gender" in header and ws.cell(row, col("gender")).value else None),
            "start": _date(ws.cell(row, col("session start")).value),
            "end": _date(ws.cell(row, col("session end")).value),
            "albums": [str(ws.cell(row, c).value).strip()
                       for c in albums if ws.cell(row, c).value],
            "scores": rated,
            "average": round(sum(rated.values()) / len(rated), 2) if rated else None,
            "stdev": round(stdev(list(rated.values())), 3) if len(rated) > 1 else None,
            "complete": len(rated) == len(members),
            "has_report": report_path(slug).exists(),
        })
    return out


def find_current(ws, header, header_row, members, aliases) -> dict:
    """The one green-filled row. Raises if there isn't exactly one."""
    rows = sorted({
        row
        for row in range(1, ws.max_row + 1)
        for cidx in range(1, ws.max_column + 1)
        if cell_is_green(ws.cell(row, cidx))
    })
    if not rows:
        raise SheetError(
            "No green-highlighted row found. The current artist is marked by "
            f"filling a row with {GREEN[2:]} (bright green).",
            kind="none")
    if len(rows) > 1:
        raise SheetError(
            f"Expected one green row, found {len(rows)} (rows {rows}). "
            "Clear the extras so the current artist is unambiguous.",
            kind="multiple", rows=rows)
    row = rows[0]
    for session in read_sessions(ws, header, header_row, members, aliases):
        if session["row"] == row:
            return session
    raise SheetError(f"Green row {row} has no artist in the Artist column.",
                     kind="unnamed", rows=[row])


# --------------------------------------------------------------------------
# Statistics -- all recomputed from raw scores
# --------------------------------------------------------------------------

def stdev(values: list[float]) -> float:
    """Sample standard deviation, matching the sheet's STDEV."""
    n = len(values)
    mean = sum(values) / n
    return math.sqrt(sum((v - mean) ** 2 for v in values) / (n - 1))


def correlation(xs: list[float], ys: list[float]) -> float | None:
    n = len(xs)
    if n < 3:
        return None
    mx, my = sum(xs) / n, sum(ys) / n
    num = sum((x - mx) * (y - my) for x, y in zip(xs, ys))
    dx = math.sqrt(sum((x - mx) ** 2 for x in xs))
    dy = math.sqrt(sum((y - my) ** 2 for y in ys))
    return round(num / (dx * dy), 3) if dx and dy else None


def compute_stats(sessions: list[dict], members: list[str]) -> dict:
    # Every fully scored session counts toward member statistics, dated or not
    # (one early session was never given a date). Only the timeline needs dates.
    rated = [s for s in sessions if s["complete"]]
    dated = sorted((s for s in rated if s["start"]), key=lambda s: s["start"])

    # Curator score: how the OTHER members rate the picks you brought.
    curator = {}
    for member in members:
        theirs = [s for s in rated if s["submitter"] == member]
        others = [s["scores"][o] for s in theirs for o in members if o != member]
        own = [s["scores"][member] for s in theirs]
        curator[member] = {
            "picks": len(theirs),
            "curator_score": round(sum(others) / len(others), 2) if others else None,
            "self_score": round(sum(own) / len(own), 2) if own else None,
            "gap": round(sum(own) / len(own) - sum(others) / len(others), 2)
                   if own and others else None,
        }

    # Taste correlation between each pair of members.
    taste = {}
    for a in members:
        for b in members:
            if a >= b:
                continue
            xs = [s["scores"][a] for s in rated]
            ys = [s["scores"][b] for s in rated]
            taste[f"{a}|{b}"] = correlation(xs, ys)

    # Contrarian index: mean distance from the other members' consensus.
    contrarian = {}
    for member in members:
        gaps = []
        for s in rated:
            peers = [s["scores"][o] for o in members if o != member]
            if peers:
                gaps.append(abs(s["scores"][member] - sum(peers) / len(peers)))
        contrarian[member] = round(sum(gaps) / len(gaps), 2) if gaps else None

    # Group average by genre.
    genres = {}
    for s in rated:
        genres.setdefault(s["genre"] or "Unlisted", []).append(s["average"])
    genre_stats = sorted(
        ({"genre": g, "sessions": len(v), "average": round(sum(v) / len(v), 2)}
         for g, v in genres.items()),
        key=lambda d: d["average"], reverse=True,
    )

    # Rolling 6-session average -- the sheet's own version of this is #REF!.
    timeline = []
    for i, s in enumerate(dated):
        window = [x["average"] for x in dated[max(0, i - 5):i + 1]]
        timeline.append({
            "date": s["start"],
            "artist": s["artist"],
            "average": s["average"],
            "rolling6": round(sum(window) / len(window), 2),
        })

    ranked = sorted(rated, key=lambda s: s["average"], reverse=True)
    return {
        "members": members,
        "sessions_rated": len(rated),
        "sessions_dated": len(dated),
        "sessions_listed": len(sessions),
        "reports": sum(1 for s in sessions if s["has_report"]),
        "overall_average": round(sum(s["average"] for s in rated) / len(rated), 2)
                           if rated else None,
        "curator": curator,
        "taste": taste,
        "contrarian": contrarian,
        "genres": genre_stats,
        "timeline": timeline,
        "ranked": [
            {k: s[k] for k in
             ("artist", "slug", "submitter", "genre", "start", "average", "stdev",
              "scores", "has_report", "albums")}
            for s in ranked
        ],
        "backlog": [
            {"artist": s["artist"], "slug": s["slug"], "submitter": s["submitter"]}
            for s in sessions if not s["complete"]
        ],
    }


# --------------------------------------------------------------------------
# CLI
# --------------------------------------------------------------------------

def open_sheet(args):
    wb = load_workbook(args.sheet_id or os.environ.get("CLUB_SHEET_ID"), args.file)
    ws = sessions_tab(wb)
    header_row, header = find_header(ws)
    return ws, header, header_row, find_members(header), load_aliases()


def emit_github_output(pairs: dict) -> None:
    path = os.environ.get("GITHUB_OUTPUT")
    if not path:
        return
    with open(path, "a", encoding="utf-8") as fh:
        for key, value in pairs.items():
            flat = " ".join(str(value).split())   # GITHUB_OUTPUT is line-based
            fh.write(f"{key}={flat}\n")


def cmd_current(args):
    ws, header, header_row, members, aliases = open_sheet(args)
    print(json.dumps(find_current(ws, header, header_row, members, aliases), indent=2))
    return 0


def cmd_check(args):
    """Decide whether a report needs generating, and say so on stdout.

    No green row is a routine between-sessions state, so it reports "nothing to
    do" and exits cleanly -- a nightly schedule shouldn't email anyone about it.
    Two green rows is a real mistake and still fails loudly, because guessing
    could write a report for the wrong artist.
    """
    ws, header, header_row, members, aliases = open_sheet(args)
    try:
        current = find_current(ws, header, header_row, members, aliases)
    except SheetError as exc:
        if exc.kind == "none":
            print("No artist is marked green in the spreadsheet yet -- "
                  "nothing to write. Highlight a row to set the next session.")
            emit_github_output({"needed": "false", "artist": "", "slug": "", "albums": ""})
            return 0
        raise

    needed = not current["has_report"]
    print(f"Current artist : {current['artist']}  (sheet row {current['row']})")
    print(f"Session        : {current['start']} -> {current['end']}")
    print(f"Report slug    : {current['slug']}")
    print(f"Report file    : html/{current['slug']}_discography.html")
    print(f"Albums         : {', '.join(current['albums']) or '(none listed)'}")
    print(f"Needs writing  : {'YES' if needed else 'no, already in the repo'}")
    emit_github_output({
        "needed": str(needed).lower(),
        "artist": current["artist"],
        "slug": current["slug"],
        "albums": "; ".join(current["albums"]),
    })
    return 0


def cmd_stats(args):
    ws, header, header_row, members, aliases = open_sheet(args)
    sessions = read_sessions(ws, header, header_row, members, aliases)
    stats = compute_stats(sessions, members)
    # A missing or ambiguous green row shouldn't block the stats refresh.
    try:
        current = find_current(ws, header, header_row, members, aliases)
        stats["current"] = {k: current[k] for k in
                            ("artist", "slug", "submitter", "start", "end",
                             "albums", "has_report")}
    except SheetError as exc:
        stats["current"] = None
        stats["current_issue"] = {"kind": exc.kind or "error",
                                  "rows": exc.rows,
                                  "message": str(exc)}
        print(f"note: no current session ({exc})", file=sys.stderr)
    else:
        stats["current_issue"] = None
    out = Path(args.out) if args.out else DATA_FILE
    out.parent.mkdir(parents=True, exist_ok=True)
    with open(out, "w", encoding="utf-8") as fh:
        json.dump(stats, fh, indent=1, ensure_ascii=False)
        fh.write("\n")
    try:
        shown = out.relative_to(REPO)
    except ValueError:
        shown = out
    print(f"Wrote {shown} -- {stats['sessions_rated']} rated "
          f"sessions, {stats['reports']} with reports.")
    return 0


def cmd_audit(args):
    """Every artist in the sheet, its slug, and whether a report exists."""
    ws, header, header_row, members, aliases = open_sheet(args)
    sessions = read_sessions(ws, header, header_row, members, aliases)
    unmapped = []
    for s in sessions:
        mark = "OK " if s["has_report"] else "   "
        print(f"{mark} r{s['row']:<4} {s['artist_raw']!r:45s} -> {s['slug']}")
        if s["artist_raw"] != s["artist"] and s["artist_raw"] not in aliases:
            unmapped.append(s["artist_raw"])
    orphans = sorted(
        p.name for p in REPORT_DIR.glob("*_discography.html")
        if p.name.replace("_discography.html", "")
        not in {s["slug"] for s in sessions}
    )
    print(f"\n{sum(1 for s in sessions if s['has_report'])}/{len(sessions)} "
          f"artists have reports.")
    if orphans:
        print(f"Reports with no matching sheet row: {', '.join(orphans)}")
    if unmapped:
        print(f"Names cleaned only by slugify (consider an alias): {unmapped}")
    return 0


def main(argv=None):
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--sheet-id", help="Google Sheets doc id (else $CLUB_SHEET_ID)")
    parser.add_argument("--file", help="read a local .xlsx instead of downloading")
    sub = parser.add_subparsers(dest="command", required=True)
    sub.add_parser("current", help="print the green-highlighted session as JSON")
    sub.add_parser("check", help="report whether the current artist needs a writeup")
    stats = sub.add_parser("stats", help="write data/ratings.json")
    stats.add_argument("--out", help="write somewhere other than data/ratings.json")
    sub.add_parser("audit", help="list every artist, its slug, and report status")

    args = parser.parse_args(argv)
    handler = {"current": cmd_current, "check": cmd_check,
               "stats": cmd_stats, "audit": cmd_audit}[args.command]
    try:
        return handler(args)
    except SheetError as exc:
        print(f"error: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    sys.exit(main())
